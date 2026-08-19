import asyncio, pathlib, tempfile, os
from playwright.async_api import async_playwright
import openpyxl

async def main():
    out = tempfile.mkdtemp()
    async with async_playwright() as p:
        b = await p.chromium.launch(executable_path="/opt/pw-browsers/chromium-1194/chrome-linux/chrome")
        pg = await b.new_page(viewport={"width":1300,"height":1000}, accept_downloads=True)
        errs=[]; pg.on("pageerror", lambda e: errs.append(str(e)))
        await pg.goto("http://localhost:8899/index.html")
        # real user path: set files on the hidden picker via the drop handlers
        await pg.set_input_files("#picker", "testdata/attendance.xlsx")
        await pg.evaluate("document.getElementById('picker').onchange({})") if False else None
        # the picker's onchange is assigned on click, so drive intake directly
        await pg.evaluate("""async () => {
            const mk = async (u, which) => {
                const blob = await (await fetch(u)).blob();
                const f = new File([blob], u.split('/').pop(), {type: blob.type});
                const dt = new DataTransfer(); dt.items.add(f);
                const node = which === 'attendance' ? document.getElementById('dropAtt') : document.getElementById('dropStu');
                node.dispatchEvent(new DragEvent('drop', {dataTransfer: dt, bubbles: true}));
            };
            await mk('testdata/attendance.xlsx','attendance');
            await mk('testdata/students.xlsx','roster');
        }""")
        await pg.wait_for_function("() => !document.getElementById('run').disabled", timeout=15000)
        await pg.click("#run")
        await pg.wait_for_selector("#results:not([hidden])")
        cards = await pg.eval_on_selector_all(".card", "els => els.map(e => e.querySelector('.n').textContent + ' ' + e.querySelector('.k').textContent)")
        print("cards:", cards)
        print("stamp:", await pg.inner_text("#stamp"))
        tabs = await pg.eval_on_selector_all(".tab", "els => els.map(e => e.textContent.trim())")
        print("tabs:", tabs)
        rows = await pg.eval_on_selector_all("tr.head", "els => els.length")
        print("rows in first tab:", rows)
        await pg.click("tr.head")
        det = await pg.eval_on_selector_all("tr.detail:not([hidden])", "els => els.length")
        print("expanded detail rows:", det)
        await pg.screenshot(path=os.path.join(out,"page.png"), full_page=False)
        # switch tab, filter
        await pg.click('.tab[data-tab="never"]')
        print("never-attended rows:", await pg.eval_on_selector_all("#tableWrap tbody tr", "els => els.length"))
        await pg.click('.tab[data-tab="owing"]')
        await pg.fill("#find", "a")
        print("filtered rows:", await pg.eval_on_selector_all("tr.head", "els => els.length"))
        await pg.fill("#find", "")
        async with pg.expect_download() as dl:
            await pg.click("#download")
        d = await dl.value
        path = os.path.join(out, d.suggested_filename)
        await d.save_as(path)
        print("downloaded:", d.suggested_filename, os.path.getsize(path), "bytes")
        print("page errors:", errs or "none")
        await b.close()
    wb = openpyxl.load_workbook(path)
    print("\nworkbook opens in openpyxl. sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  {name:<16} {ws.max_row-1:>5} data rows x {ws.max_column} cols | header: {[c.value for c in ws[1]][:4]}")
    print(os.path.join(out,'page.png'))

asyncio.run(main())
